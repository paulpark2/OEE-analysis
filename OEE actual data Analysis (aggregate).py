#!/usr/bin/env python
# coding: utf-8

# FP	Solder die attachLSD003, LSD004, LSD005, LSD006, LSD007, LSD009, LSD010, LSD013, LSD014, LSD015, LSD016, LSD018, LSD019, LSD024, LSD025, LSD026, LSD036, LSD037
# FP	AL wire bonding	LAW004, LAW006, LAW007, LAW008, LAW009, LAW010, LAW011, LAW012, LAW013, LAW016, LAW018, LAW020, LAW022, LAW024, LAW026, LAW030, LAW033, LAW034, LAW035, LAW036, LAW040, LAW043, LAW050
# FP	Mold - ASM	LMD002, LMD003, LMD004, LMD005, LMD006
# FP	Mold - Hanmi	LMD007, LMD011
# DCB	Solder die attach	LSD001, LSD002, LSD011,LSD017, LSD008, LSD012, LSD011, LSD017, LSD022, LSD023, LSD027, LSD028, LSD029, LSD035, LSD038, LSD039, LSD040, LSD042
# DCB	AL wire bonding - Power & gate	LAW028, LAW029, LAW032, LAW038, LAW045, LAW046, LAW053
# DCB	AL wire bonding - Signal	LAW014, LAW037, LAW042, LAW044, LAW052
# DCB	Mold - Hanmi	LMD009, LMD011, LMD012
# Mini Common	PCB singulation	LPS002
# Mini Common	PCB connection	LPC001, LPC002
# Mini Common	PCB singualtion & connection	LPC003, LPC004, LPC005
# Mini Common	Epoxy die attach	LED001, LED002, LED003, LED004, LED005, LED006, LED007, LED008, LED011, LED012, LED013
# Mini Common	Au wire bonding	LGW001, LGW002, LGW003, LGW004, LGW005, LGW006, LGW007, LGW008, LGW009, LGW010, LGW012, LGW013, LGW014, LGW015, LGW018, LGW019, LGW020, LGW021
# Mini Common	IVI	AIVI001, AIVI002, AIVI003, AIVI004, AIVI005, AIVI006
# Mini Common	Laser marking	LMK001, LMK002, LMK003, LMK005
# Mini Common	Trim & form	LTF001, LTF002, LTF003
# Mini Common	Test Handler system	LTE002, LTE004, LTE005, LTE006, LTE007, LTE010, LTE011, LTE012, LTE013, LTE014
# Mini Common	AFVI	AFVI001, AFVI002
# Tiny	Solder die attach	LSD020, LSD021, LSD030, LSD031
# Tiny	Epoxy die attach	LED009, LED010
# Tiny	Potting	LED015, LED018, LED019
# Tiny	Al wire bonding -1x head	LAW041, LAW047
# Tiny	Al wire bonding -3x head	LAW031, LAW039, LAW049
# Tiny	Plasma - Au WB	LPL002
# Tiny	Au wire bonding	LGW011, LGW016, LGW017
# Tiny	IVI	AIVI007, AIVI008
# Tiny	Plasma - Mold	LPL003
# Tiny	Mold - ASM	LMD010, LMD013
# Tiny	Laser marking	LMK004
# Tiny	Trim & form	LTF005
# Tiny	Leadcoversion (Tiny)	LLC001, LLC002
# Tiny	Test Handler system	LTE008, LTE009
# Tiny	AFVI	AFVI003
# Tiny 3.0	Solder die attach	LSD041
# Tiny 3.0	PCB singulation	LPS002
# Tiny 3.0	PCB connection	LPC007
# Tiny 3.0	Epoxy die attach	LED020, LED014
# Tiny 3.0	AL wire bonding - Power & gate	LAW048
# Tiny 3.0	AL wire bonding - Signal	LAW051
# Tiny 3.0	Au wire bonding	LGW023
# Tiny 3.0	IVI	AIVI009
# Tiny 3.0	Plasma - Mold	LPL004
# Tiny 3.0	Mold - ASM	LMD014
# Tiny 3.0	Laser marking	LMK004
# Tiny 3.0	Trim & form	LTF008
# Tiny 3.0	Test Handler system	LTE015
# Tiny 3.0	AFVI	AFVI003

# PFC0 = ['LSD001','LSD012','LSD012','LSD023','LSD028','LSD035','LSD039','LSD040']
# PFC4 = ['LSD001','LSD012','LSD023','LSD027','LSD035','LSD039','LSD040']
# PFC8 = ['LSD001','LSD012','LSD023','LSD027','LSD035','LSD038']
# 
# <All SDA>
# 'LSD001', 'LSD002', 'LSD003', 'LSD004', 'LSD005', 'LSD006', 'LSD007', 'LSD008', 'LSD009', 'LSD010', 'LSD011', 'LSD012', 'LSD013', 'LSD014', 'LSD015', 'LSD016', 'LSD017', 'LSD018', 'LSD019', 'LSD020', 'LSD021', 'LSD022', 'LSD023', 'LSD024', 'LSD025', 'LSD026', 'LSD027', 'LSD028', 'LSD029', 'LSD030', 'LSD031', 'LSD035', 'LSD036', 'LSD037', 'LSD038', 'LSD039', 'LSD040', 'LSD041', 'LSD042'
# 
# <BN SDA>
# 'LSD003', 'LSD004', 'LSD005', 'LSD006', 'LSD007', 'LSD009', 'LSD010', 'LSD013', 'LSD014', 'LSD015', 'LSD016', 'LSD018', 'LSD019', 'LSD024', 'LSD025', 'LSD026', 'LSD036', 'LSD037','LSD001', 'LSD011', 'LSD012', 'LSD017', 'LSD023', 'LSD027', 'LSD028', 'LSD029', 'LSD035', 'LSD038', 'LSD039', 'LSD040','LSD020', 'LSD021', 'LSD030', 'LSD031','LSD041'
PFC0 = ['LSD002','LSD011','LSD027','LSD022','LSD008','LSD038','LSD029','LSD042']
PFC4 = ['LSD002','LSD011','LSD017','LSD028','LSD022','LSD008','LSD038','LSD029','LSD042']
PFC8 = ['LSD002','LSD011','LSD017','LSD028','LSD022','LSD008','LSD040','LSD029','LSD039','LSD042']
df = df[~(df['Month'].isin(['2022-03']) & df['EQ_ID'].isin(PFC4))]
df = df[~(df['LOG_WEEK'].isin(['LW-2023-06']) & df['EQ_ID'].isin(PFC4))]
df = df[~(df['LOG_WEEK'].isin(['LW-2023-07']) & df['EQ_ID'].isin(PFC4))]
df = df[~(df['LOG_WEEK'].isin(['LW-2023-08']) & df['EQ_ID'].isin(PFC0))]
# In[1]:


import numpy as np
import pandas as pd
import csv
import openpyxl
from openpyxl.utils import range_boundaries
import os
from openpyxl import load_workbook


# In[2]:


df_list = [
    #pd.read_csv('202010-202105OEE.csv'),
    #pd.read_csv('202106-202111OEE.csv'),
    #pd.read_csv('202112-2201OEE.csv'),
    #pd.read_csv('202112OEE.csv'),
    #pd.read_csv('202201OEE.csv'),
    #pd.read_csv('202202OEE.csv'),
    #pd.read_csv('202203OEE.csv'),
    #pd.read_csv('202204OEE.csv'),
    #pd.read_csv('202205OEE.csv'),
    #pd.read_csv('202206OEE.csv'),
    pd.read_csv('202207OEE.csv'),
    pd.read_csv('202208OEE.csv'),
    pd.read_csv('202209OEE.csv'),
    pd.read_csv('202210OEE.csv'),
    pd.read_csv('202211OEE.csv'),
    pd.read_csv('202212OEE.csv'),
    pd.read_csv('202301OEE.csv'),
    pd.read_csv('202302OEE.csv'),
    pd.read_csv('202303OEE.csv')
]


# In[3]:


df = pd.concat(df_list)


# In[4]:


df = df[['LOG_WEEK', 'LOG_DATE', 'EQUIPMENT_NAME', 'LEVEL5_NAME', 'LEVEL6_NAME', 'DURATION_HOUR','PACKAGE_CLASS_PD', 'PPOS']]


# In[5]:


df.rename(columns = {'EQUIPMENT_NAME' : 'EQ_ID'}, inplace = True)


# In[6]:


#df.rename(columns = {'Log Week' : 'LOG_WEEK', 'Fiscal Date' : 'LOG_DATE', 'Equipment Name' : 'EQ_ID', 'Level5 Name' : 'LEVEL5_NAME','Level6 Name' : 'LEVEL6_NAME', 'Duration Hour' : 'DURATION_HOUR'}, inplace = True)
#df0.rename(columns = {'Log Week' : 'LOG_WEEK', 'Fiscal Date' : 'LOG_DATE', 'Equipment Name' : 'EQ_ID', 'Level5 Name' : 'LEVEL5_NAME','Level6 Name' : 'LEVEL6_NAME', 'Duration Hour' : 'DURATION_HOUR'}, inplace = True)


# In[6]:


df = df[(df.LEVEL5_NAME != 'NO MATERIAL')]
#df = df[(df.LEVEL5_NAME != 'SPEED LOSS')]
df = df[(df.LEVEL5_NAME != 'INVALID DATA')]
df = df[(df.LEVEL5_NAME != 'EXCEPTION')]
df = df[(df.LEVEL5_NAME != 'NO COMMUNICATION')]
df = df[(df.LEVEL5_NAME != 'NON-SCHEDULED TIME')]
#df = df[(df.LEVEL5_NAME != 'ENGINEERING EVALUATION')]
#df = df[(df.LEVEL5_NAME != '<unrecognized>')]
df = df[(df.EQ_ID != 'LSD999')]
df = df[(df.EQ_ID != 'LAW099')]
df = df[(df.EQ_ID != 'LSD099')]
df = df[(df.EQ_ID != 'AFVI099')]
df = df[(df.LOG_DATE != '2022-08-29') | (df.EQ_ID != 'LAW004') | (df.LEVEL6_NAME != 'WAITING FOR RESPONSE')]
df = df[(df.LOG_DATE != '2022-08-27') | (df.EQ_ID != 'LAW004') | (df.LEVEL6_NAME != 'WAITING FOR RESPONSE')]
df = df[(df.LOG_DATE != '2022-08-28') | (df.EQ_ID != 'LAW004') | (df.LEVEL6_NAME != 'WAITING FOR RESPONSE')]
#LAW004 2022-08-27  5:48:00 오전 ~ 2022-08-29  8:42:00 오전 Bitlocker  (Waiting for response)
df = df[(df.LOG_DATE != '2022-08-28') | (df.EQ_ID != 'LMD012') | (df.LEVEL6_NAME != 'WAITING FOR RESPONSE') | (df.PACKAGE_CLASS_PD != 'MDIP-DCB')]
df = df[(df.LOG_DATE != '2022-08-29') | (df.EQ_ID != 'LMD012') | (df.LEVEL6_NAME != 'WAITING FOR RESPONSE') | (df.PACKAGE_CLASS_PD != 'MDIP-DCB')]
#LMD012 2022-08-28  4:17:00 오전 ~ 2022-08-29  8:45:00 오전 Bitlocker  (Waiting for response)
df.drop(df.loc[(df.LOG_DATE == '2022-10-22') & (df.EQ_ID == 'LMD011') & (df.LEVEL6_NAME == 'WAITING FOR RESPONSE') & (df.DURATION_HOUR > 4)].index,inplace=True)
#LMD011 2022-10-22 9:35:00 오전 ~ 2022-10-22  6:25:00 오후 miss key-in  (Waiting for response)
df = df[(df.LOG_DATE != '2022-10-25') | (df.EQ_ID != 'LSD018') | (df.LEVEL6_NAME != 'CHANGE COLLET')]
df = df[(df.LOG_DATE != '2022-10-25') | (df.EQ_ID != 'LSD019') | (df.LEVEL6_NAME != 'CHANGE COLLET')]
df = df[(df.LOG_DATE != '2022-10-18') | (df.EQ_ID != 'LSD007') | (df.LEVEL6_NAME != 'CHANGE COLLET')]
df = df[(df.LOG_DATE != '2022-10-18') | (df.EQ_ID != 'LSD006') | (df.LEVEL6_NAME != 'CHANGE COLLET')]

mask = (df.LOG_DATE == '2023-02-13') & (df.EQ_ID == 'LMD004') & (df.LEVEL6_NAME == 'DAILY, WEEKLY')
df.loc[mask, 'LEVEL5_NAME'] = 'MACHINE FAILURE'
df.loc[mask, 'LEVEL6_NAME'] = 'INPUT HANDLER PROBLEM'
mask2 = (df.LOG_DATE == '2023-02-15') & (df.EQ_ID == 'LMD004') & (df.LEVEL6_NAME == 'DAILY, WEEKLY')
df.loc[mask2, 'LEVEL5_NAME'] = 'MACHINE FAILURE'
df.loc[mask2, 'LEVEL6_NAME'] = 'PELLET PROBLEM'

#df = df[(df.EQ_ID != 'LSD032')]
#df = df[(df.EQ_ID != 'LSD033')]


# In[7]:


df['Process'] = df['EQ_ID'].str.extract(pat = '([A-Z]..)')


# In[8]:


df['Month'] = pd.to_datetime(df['LOG_DATE']).dt.to_period('M') #.astype(str)


# In[10]:


#df.loc[df['PACKAGE_CLASS_PD'].isin(['MDIP-FP', 'MDIP-DCB']), 'EQ_ID'] = 'new_value'


# In[9]:


#PFC bottleneck 기준 data 뽑기 (22년 10월부터 가능)
PFC_D = ['IM564-X6D','IFCM30T65GD','IFCM20T65GD', 'IFCM15S60GS', 'IFCM15P60GD', 'IFCM15S60GD']
PFC_EQ = ['LSD011', 'LSD017','LSD028', 'LSD029', 'LSD040', 'LSD039']
PFC_EQ2 = ['LSD011', 'LSD027','LSD029', 'LSD038']
df.sort_values(by=['EQ_ID','LOG_DATE'], inplace=True)
df['PPOS'] = np.where((df.Month >= '2021-12'), df['PPOS'].fillna(method = 'bfill'), df['PPOS'])
df['PPOS'] = np.where((df.Month >= '2021-12'), df['PPOS'].fillna(method = 'ffill'), df['PPOS'])

df = df[~( (df['PPOS'].isin(PFC_D)) & (df['EQ_ID'].isin(PFC_EQ)) & (df.Month >= '2021-12')) ]
df = df[~( (~df['PPOS'].isin(PFC_D)) & (df['EQ_ID'].isin(PFC_EQ2)) & (df.Month >= '2021-12')) ]

#df = df[~(np.logical_not(df['PPOS'].isin(PFC_D)) & df['EQ_ID'].isin(PFC_EQ2) & df['Month'].isin(['2022-10','2022-11']) )]


# In[11]:


#df.loc[df['PACKAGE_CLASS_PD'] == 'MDIP-FP', 'EQ_ID'] = 'new_value'


# In[10]:


pq = pd.read_csv('EQ_Process_BN.csv')


# In[11]:


pq['EQ_ID'] = pq['EQ_ID'].str.strip()


# In[12]:


target = pd.read_csv('OEE Target3.csv')


# In[13]:


target['Level 5'] = target['Level 5'].str.upper()


# In[16]:


#df['week'] = pd.to_datetime(df['LOG_DATE']).dt.to_period('Y').astype(str) + '_' + df['LOG_WEEK'].astype(str)


# In[22]:


com = {
    #BN SDA
    #'a' : ['LSD003', 'LSD004', 'LSD005', 'LSD006', 'LSD007', 'LSD009', 'LSD010', 'LSD013', 'LSD014', 'LSD015', 'LSD016', 'LSD018', 'LSD019', 'LSD024', 'LSD025', 'LSD026', 'LSD036', 'LSD037','LSD001', 'LSD011', 'LSD012', 'LSD017', 'LSD023', 'LSD027', 'LSD028', 'LSD029', 'LSD035', 'LSD038', 'LSD039', 'LSD040','LSD020', 'LSD021', 'LSD030', 'LSD031','LSD041']
    'b' : ['LMD002', 'LMD003', 'LMD004', 'LMD005', 'LMD006', 'LMD007', 'LMD009', 'LMD010', 'LMD011', 'LMD012', 'LMD013', 'LMD14']
    #'a' : ['LAW004', 'LAW006', 'LAW007', 'LAW008', 'LAW009', 'LAW010', 'LAW011', 'LAW012', 'LAW013', 'LAW016', 'LAW018', 'LAW020', 'LAW022', 'LAW024', 'LAW026', 'LAW030', 'LAW033', 'LAW034', 'LAW035', 'LAW036', 'LAW040', 'LAW043', 'LAW050']
    #'b' : ['LSD001', 'LSD012', 'LSD023', 'LSD027', 'LSD035', 'LSD038']
    #'c': ['LAW004','LAW005', 'LAW006', 'LAW007', 'LAW008', 'LAW009', 'LAW010', 'LAW011', 'LAW012', 'LAW013', 'LAW016', 'LAW018', 'LAW020', 'LAW022', 'LAW024', 'LAW026', 'LAW030', 'LAW033', 'LAW034', 'LAW035', 'LAW036', 'LAW040', 'LAW043']
    #'d': ['LSD001', 'LSD002','LSD008','LSD011', 'LSD012', 'LSD017', 'LSD022','LSD023', 'LSD027', 'LSD028', 'LSD029', 'LSD035', 'LSD038', 'LSD039', 'LSD040','LSD042']
    #'e':['LSD001', 'LSD003', 'LSD004', 'LSD005', 'LSD006', 'LSD007', 'LSD009', 'LSD010', 'LSD012', 'LSD013', 'LSD014', 'LSD015', 'LSD016', 'LSD018', 'LSD019', 'LSD020', 'LSD021', 'LSD023', 'LSD024', 'LSD025', 'LSD026', 'LSD027', 'LSD030', 'LSD031', 'LSD035', 'LSD036', 'LSD037', 'LSD038', 'LSD041']
      }


# In[23]:


monthly(com, df, pq, target)


# In[15]:


weekly(com, df, pq, target)


# In[18]:


daily(com, df, pq, target)


# In[27]:


k = pd.crosstab([df2.LEVEL5_NAME,df2.LEVEL6_NAME],df2.Month,values=df2.DURATION_HOUR,aggfunc=sum,rownames=['Level 5','LV6'],colnames=['month'], normalize='columns')


# In[31]:


k.to_excel("ch.xlsx")


# In[18]:


ALL(df, pq, target) # LV6 of all process and each file


# In[33]:


one_page(df, pq, target) # LV5 OF all process and one file


# In[24]:


lst = ['LSD003', 'LSD004', 'LSD005', 'LSD006', 'LSD007', 'LSD009', 'LSD010', 'LSD013', 'LSD014', 'LSD015', 'LSD016', 'LSD018', 'LSD019', 'LSD024', 'LSD025', 'LSD026', 'LSD036', 'LSD037','LSD001', 'LSD011', 'LSD012', 'LSD017', 'LSD023', 'LSD027', 'LSD028', 'LSD029', 'LSD035', 'LSD038', 'LSD039', 'LSD040','LSD020', 'LSD021', 'LSD030', 'LSD031','LSD041']


# In[22]:


lst = ['LSD003', 'LSD004', 'LSD005', 'LSD006', 'LSD007', 'LSD009', 'LSD010', 'LSD013', 'LSD014', 'LSD015', 'LSD016', 'LSD018', 'LSD019', 'LSD024', 'LSD025', 'LSD026', 'LSD036', 'LSD037','LSD001', 'LSD011', 'LSD012', 'LSD017', 'LSD023', 'LSD027', 'LSD028', 'LSD029', 'LSD035', 'LSD038', 'LSD039', 'LSD040','LSD020', 'LSD021', 'LSD030', 'LSD031','LSD041']


# In[25]:


filter1 = df['EQ_ID'].isin(lst)
df2 = df[filter1]


# In[26]:


df2 = df2[(df2.Month == '2023-03')]


# In[33]:


EQ_ID(df2, lst, pq, target) #LV5


# # Monthly loss analysis

# In[20]:


def monthly(com, df, pq, target):
        
    def create_merged_cell_lookup(sheet) -> dict:
        merged_lookup = {}
        for cell_group in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
            if min_col == max_col:
                top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
                merged_lookup[str(cell_group)] = top_left_cell_value
        return merged_lookup    
    
    for x,y in com.items():
        filter1 = df['EQ_ID'].isin(y)
        df2 = df[filter1]
        #df = df[(df.LEVEL5_NAME != 'NO MATERIAL')]
        #if x == 'a': df2 = df2[(df2.Month == '2022-05') | (df2.Month == '2022-06') | (df2.Month == '2022-07')]
        #if x == 'a': df2 =  df2.sort_values(by=['LOG_DATE']).fillna(method='ffill')[(df2.PACKAGE_CLASS_PD == 'MDIP-FP')]
        #df2 = df2.groupby(['Month','Process','LEVEL5_NAME','LEVEL6_NAME'], as_index=False).sum()
        globals()[x] =pd.crosstab([df2.LEVEL5_NAME,df2.LEVEL6_NAME],df2.Month,values=df2.DURATION_HOUR,aggfunc=sum,rownames=['Level 5','LV6'],colnames=['month'], normalize='columns')
        #globals()[x] = pd.crosstab([df2.LEVEL5_NAME],df2.LOG_WEEK,values=df2.DURATION_HOUR,aggfunc=sum,rownames=['Level 5'],colnames=['date'], normalize='columns')

        try:
            p = globals()[x].loc[['NORMAL PRODUCTION','SPEED LOSS']].sum()
            p.name = 'NORMAL PRODUCTION'
            p = pd.DataFrame(p)
            p = p.transpose()
            p.insert(0, "LV6", ['NORMAL PRODUCTION'], True)
            p = p.set_index('LV6',append=True)
            globals()[x] = globals()[x].drop(['NORMAL PRODUCTION','SPEED LOSS'])
            globals()[x] = pd.concat([globals()[x], p])
            #globals()[x] = globals()[x].append(p)
        except KeyError:
            pass    

        k = pq[pq['EQ_ID'].str.contains(com.get(x)[0])]
        tar = target[(target.Device == k.iloc[0]['Device']) & (target.Process == k.iloc[0]['Process'])]
        globals()[x] = globals()[x].join(tar.set_index('Level 5')['Loss factor'], on='Level 5')
        globals()[x].to_excel(k.iloc[0]['Device'] + '_' + k.iloc[0]['Process'] + '.xlsx')

        wbook = openpyxl.load_workbook(k.iloc[0]['Device'] + '_' + k.iloc[0]['Process'] + '.xlsx')
        sheet = wbook["Sheet1"]
        lookup = create_merged_cell_lookup(sheet)
        cell_group_list = lookup.keys()
        for cell_group in cell_group_list:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
            sheet.unmerge_cells(str(cell_group))
            for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = lookup[cell_group]
        wbook.save(k.iloc[0]['Device'] + '_' + k.iloc[0]['Process'] + '.xlsx')


# # Weekly loss analysis

# In[14]:


def weekly(com, df, pq, target):

    def create_merged_cell_lookup(sheet) -> dict:
        merged_lookup = {}
        for cell_group in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
            if min_col == max_col:
                top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
                merged_lookup[str(cell_group)] = top_left_cell_value
        return merged_lookup        
    
    for x,y in com.items():
        filter1 = df['EQ_ID'].isin(y)
        df2 = df[filter1]
        #df = df[(df.LEVEL5_NAME != 'NO MATERIAL')]
        #if x == 'a': df2 = df2[(df2.Month == '2022-05') | (df2.Month == '2022-06') | (df2.Month == '2022-07')]
        #if x == 'a': df2 = df2[(df2.PACKAGE_CLASS_PD == 'MDIP-FP')]    
        df2 = df2.groupby(['LOG_WEEK','Process','LEVEL5_NAME','LEVEL6_NAME'], as_index=False).sum()
        globals()[x] =pd.crosstab([df2.LEVEL5_NAME,df2.LEVEL6_NAME],df2.LOG_WEEK,values=df2.DURATION_HOUR,aggfunc=sum,rownames=['Level 5','LV6'],colnames=['month'], normalize='columns')
        #globals()[x] = pd.crosstab([df2.LEVEL5_NAME],df2.LOG_WEEK,values=df2.DURATION_HOUR,aggfunc=sum,rownames=['Level 5'],colnames=['date'], normalize='columns')

        try:
            p = globals()[x].loc[['NORMAL PRODUCTION','SPEED LOSS']].sum()
            p.name = 'NORMAL PRODUCTION'
            p = pd.DataFrame(p)
            p = p.transpose()
            p.insert(0, "LV6", ['NORMAL PRODUCTION'], True)
            p = p.set_index('LV6',append=True)
            globals()[x] = globals()[x].drop(['NORMAL PRODUCTION','SPEED LOSS'])
            globals()[x] = pd.concat([globals()[x], p])
            #globals()[x] = globals()[x].append(p)
        except KeyError:
            pass

        k = pq[pq['EQ_ID'].str.contains(com.get(x)[0])]
        tar = target[(target.Device == k.iloc[0]['Device']) & (target.Process == k.iloc[0]['Process'])]
        globals()[x] = globals()[x].join(tar.set_index('Level 5')['Loss factor'], on='Level 5')
        globals()[x].to_excel(k.iloc[0]['Device'] + '_' + k.iloc[0]['Process'] + '.xlsx')

        wbook = openpyxl.load_workbook(k.iloc[0]['Device'] + '_' + k.iloc[0]['Process'] + '.xlsx')
        sheet = wbook["Sheet1"]
        lookup = create_merged_cell_lookup(sheet)
        cell_group_list = lookup.keys()
        for cell_group in cell_group_list:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
            sheet.unmerge_cells(str(cell_group))
            for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = lookup[cell_group]
        wbook.save(k.iloc[0]['Device'] + '_' + k.iloc[0]['Process'] + '.xlsx')


# # Daily loss analysis

# In[16]:


def daily(com, df, pq, target):

    def create_merged_cell_lookup(sheet) -> dict:
        merged_lookup = {}
        for cell_group in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
            if min_col == max_col:
                top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
                merged_lookup[str(cell_group)] = top_left_cell_value
        return merged_lookup        

    for x,y in com.items():
        filter1 = df['EQ_ID'].isin(y)
        df2 = df[filter1]
        #if x == 'c': df2 = df2[(df2.Month == '2022-07')]
        #if x == 'a': df2 = df2[(df2.PACKAGE_CLASS_PD == 'MDIP-FP')]    
        df2 = df2.groupby(['LOG_DATE','Process','LEVEL5_NAME','LEVEL6_NAME'], as_index=False).sum()
        globals()[x] =pd.crosstab([df2.LEVEL5_NAME,df2.LEVEL6_NAME],df2.LOG_DATE,values=df2.DURATION_HOUR,aggfunc=sum,rownames=['Level 5','LV6'],colnames=['month'], normalize='columns')
        #globals()[x] = pd.crosstab([df2.LEVEL5_NAME],df2.LOG_WEEK,values=df2.DURATION_HOUR,aggfunc=sum,rownames=['Level 5'],colnames=['date'], normalize='columns')

        try:
            p = globals()[x].loc[['NORMAL PRODUCTION','SPEED LOSS']].sum()
            p.name = 'NORMAL PRODUCTION'
            p = pd.DataFrame(p)
            p = p.transpose()
            p.insert(0, "LV6", ['NORMAL PRODUCTION'], True)
            p = p.set_index('LV6',append=True)
            globals()[x] = globals()[x].drop(['NORMAL PRODUCTION','SPEED LOSS'])
            globals()[x] = pd.concat([globals()[x], p])
        except KeyError:
            pass    

        k = pq[pq['EQ_ID'].str.contains(com.get(x)[0])]
        tar = target[(target.Device == k.iloc[0]['Device']) & (target.Process == k.iloc[0]['Process'])]
        globals()[x] = globals()[x].join(tar.set_index('Level 5')['Loss factor'], on='Level 5')
        globals()[x].to_excel(k.iloc[0]['Device'] + '_' + k.iloc[0]['Process'] + '.xlsx')

        wbook = openpyxl.load_workbook(k.iloc[0]['Device'] + '_' + k.iloc[0]['Process'] + '.xlsx')
        sheet = wbook["Sheet1"]
        lookup = create_merged_cell_lookup(sheet)
        cell_group_list = lookup.keys()
        for cell_group in cell_group_list:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
            sheet.unmerge_cells(str(cell_group))
            for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = lookup[cell_group]
        wbook.save(k.iloc[0]['Device'] + '_' + k.iloc[0]['Process'] + '.xlsx')


# # All processes

# In[17]:


def ALL(df, pq, target):

    def create_merged_cell_lookup(sheet) -> dict:
        merged_lookup = {}
        for cell_group in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
            if min_col == max_col:
                top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
                merged_lookup[str(cell_group)] = top_left_cell_value
        return merged_lookup

    share = pd.DataFrame({
        'Name' : ['FPMold - Hanmi', 'DCBMold - Hanmi', 'TinyLaser marking', 'Tiny 3.0Laser marking', 'TinyAFVI', 'Tiny 3.0AFVI'],
        'PACKAGE_CLASS_PD' : ['MDIP-DCB, MDIP-DCBPLUS', 'MDIP-FP', 'MDIP-TINY3', 'MDIP-TINY, MSIP-TINY ', 'MDIP-DCB, MDIP-DCBPLUS, MDIP-FP, MDIP-TINY3', 'MDIP-DCB, MDIP-DCBPLUS, MDIP-FP, MDIP-TINY, MSIP-TINY' ]
            })
    s = share['Name'].tolist()
    
    for index, row in pq.iterrows():
        k = row
        lst = row['EQ_ID'].split(',')
        lst = [x.strip(' ') for x in lst]
        df2 = df[df['EQ_ID'].isin(lst)]

        if row['Device'] + row['Process'] in s:
            df2 = df2.sort_values(by=['LOG_DATE'])
            df2 = df2.fillna(method='ffill')
            l = share.loc[share.Name == row['Device'] + row['Process'],'PACKAGE_CLASS_PD'].tolist()
            l = " ".join(l)
            l = l.split(',')
            l = [x.strip(' ') for x in l]
            for i in l:
                df2 = df2[(df2.PACKAGE_CLASS_PD != i)]


        #df2 = df2.groupby(['Month','Process','LEVEL5_NAME','LEVEL6_NAME'], as_index=False).sum()
        df3 = pd.crosstab([df2.LEVEL5_NAME,df2.LEVEL6_NAME],df2.LOG_WEEK,values=df2.DURATION_HOUR,aggfunc=sum,rownames=['Level 5','LV6'],colnames=['month'], normalize='columns')

        try:
            p = df3.loc[['NORMAL PRODUCTION','SPEED LOSS']].sum()
            p.name = 'NORMAL PRODUCTION'
            p = pd.DataFrame(p)
            p = p.transpose()
            p.insert(0, "LV6", ['NORMAL PRODUCTION'], True)
            p = p.set_index('LV6',append=True)
            df3 = df3.drop(['NORMAL PRODUCTION','SPEED LOSS'])
            df3 = pd.concat([df3, p])
            #df3 = df3.append(p)
        except KeyError:
            pass           
        
        tar = target[(target.Device == row['Device']) & (target.Process == row['Process'])]    
        df4 = df3.join(tar.set_index('Level 5')['Loss factor'], on='Level 5')
        df4 = df4.round(4)
        df4.to_excel(row['Device'] + '_' + row['Process'] + '.xlsx')

        wbook = openpyxl.load_workbook(row['Device'] + '_' + row['Process'] + '.xlsx')
        sheet = wbook["Sheet1"]
        lookup = create_merged_cell_lookup(sheet)
        cell_group_list = lookup.keys()
        for cell_group in cell_group_list:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
            sheet.unmerge_cells(str(cell_group))
            for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
                for cell in row:
                    cell.value = lookup[cell_group]
        wbook.save(k['Device'] + '_' + k['Process'] + '.xlsx')


# # LV5 one page summary

# In[1]:


def one_page(df, pq, target):
    
    share = pd.DataFrame({
        'Name' : ['FPMold - Hanmi', 'DCBMold - Hanmi', 'TinyLaser marking', 'Tiny 3.0Laser marking', 'TinyAFVI', 'Tiny 3.0AFVI'],
        'PACKAGE_CLASS_PD' : ['MDIP-DCB, MDIP-DCBPLUS', 'MDIP-FP', 'MDIP-TINY3', 'MDIP-TINY, MSIP-TINY ', 'MDIP-DCB, MDIP-DCBPLUS, MDIP-FP, MDIP-TINY3', 'MDIP-DCB, MDIP-DCBPLUS, MDIP-FP, MDIP-TINY, MSIP-TINY' ]
            })
    s = share['Name'].tolist()

    def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                           truncate_sheet=False,
                           **to_excel_kwargs):


        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a',if_sheet_exists='overlay')
        
        try:
            FileNotFoundError
        except NameError:
            FileNotFoundError = IOError
            
        try:
            # try to open an existing workbook
            writer.book = load_workbook(filename)

            # get the last row in the existing Excel sheet
            # if it was not specified explicitly
            if startrow is None and sheet_name in writer.book.sheetnames:
                startrow = writer.book[sheet_name].max_row

            # truncate sheet
            if truncate_sheet and sheet_name in writer.book.sheetnames:
                # index of [sheet_name] sheet
                idx = writer.book.sheetnames.index(sheet_name)
                # remove [sheet_name]
                writer.book.remove(writer.book.worksheets[idx])
                # create an empty sheet [sheet_name] using old index
                writer.book.create_sheet(sheet_name, idx)

            # copy existing sheets
            writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        except FileNotFoundError:
            # file does not exist yet, we will create it
            pass
            
        if startrow is None:
            startrow = 0
        #with pd.ExcelWriter('recent_OEE3.xlsx', mode='a') as writer:
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
        
        writer.save()
        
    filename = r'C:\Users\ParkPaul\Documents\Python\recent_OEE3.xlsx'

    for index, row in pq.iterrows():
        k = row
        lst = row['EQ_ID'].split(',')
        lst = [x.strip(' ') for x in lst]
        df2 = df[df['EQ_ID'].isin(lst)]

        if row['Device'] + row['Process'] in s:
            df2 = df2.sort_values(by=['LOG_DATE'])
            df2 = df2.fillna(method='ffill')
            l = share.loc[share.Name == row['Device'] + row['Process'],'PACKAGE_CLASS_PD'].tolist()
            l = " ".join(l)
            l = l.split(',')
            l = [x.strip(' ') for x in l]
            for i in l:
                df2 = df2[(df2.PACKAGE_CLASS_PD != i)]

        df2 = pd.crosstab([df2.LEVEL5_NAME],df2.Month,values=df2.DURATION_HOUR,aggfunc=sum,rownames=['Level 5'],colnames=['month'], normalize='columns')

        try:
            p = df2.loc[['NORMAL PRODUCTION','SPEED LOSS']].sum()
            p.name = 'NORMAL PRODUCTION'
            df2 = df2.drop(['NORMAL PRODUCTION','SPEED LOSS'])
            pd.concat([df2, p.transpose()])
        except KeyError:
            pass

        df2['Average'] = df2.mean(axis=1)
        tar = target[(target.Device == row['Device']) & (target.Process == row['Process'])]
        df3 = df2.join(tar.set_index('Level 5')['Loss factor'], on='Level 5')
        df3 = df3.reset_index()
        df3 = df3.rename(columns={'Level 5': row['Device'] + "_" + row['Process']})
        df3 = df3.round(4)
        df3 = df3.sort_values(by='Loss factor', ascending=False)


        append_df_to_excel(filename, df3)


# In[18]:


one_page(df, pq, target) # LV5 OF all process and one file


# # EQ_ID

# In[32]:


def EQ_ID(df2, lst, pq, target):
    def create_merged_cell_lookup(sheet) -> dict:
        merged_lookup = {}
        for cell_group in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
            if min_col == max_col:
                top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
                merged_lookup[str(cell_group)] = top_left_cell_value
        return merged_lookup

    df3 = df2.pivot_table(index=['LEVEL5_NAME'], columns = ['EQ_ID','Month'], values='DURATION_HOUR', aggfunc={'sum'}).apply(lambda y:y/float(y.sum())).round(2)
    
    try:
        p = df3.loc[['NORMAL PRODUCTION','SPEED LOSS']].sum()
        p.name = 'NORMAL PRODUCTION'
        df3 = df3.drop(['NORMAL PRODUCTION','SPEED LOSS'])
        df3 = df3.append(p.transpose())
        #pd.concat([df3, p.transpose()])
    except KeyError:
        pass
    
    k = pq[pq['EQ_ID'].str.contains(lst[0])]
    tar = target[(target.Device == k.iloc[0]['Device']) & (target.Process == k.iloc[0]['Process'])]
    df3 = df3.reset_index()
    df3 = df3.rename(columns={'LEVEL5_NAME': 'Level 5'})
    df3 = df3.set_index('Level 5')
    tar = pd.DataFrame(tar.set_index('Level 5')['Loss factor'])
    tar.columns = pd.MultiIndex.from_product([['a'],['b'], tar.columns])
    df3 = df3.join(tar, on='Level 5')
    df3 = df3.replace(np.nan, 0)
    col = df3.columns.tolist()
    col = col[-1:] + col[:-1]
    df3 = df3[col]
    df3.sort_values(by=[('a', 'b', 'Loss factor')], inplace=True, ascending=False)
    df3.to_excel(k.iloc[0]['Device'] + '_' + k.iloc[0]['Process'] + '.xlsx')

    wbook = openpyxl.load_workbook(k.iloc[0]['Device'] + '_' + k.iloc[0]['Process'] + '.xlsx')
    sheet = wbook["Sheet1"]
    lookup = create_merged_cell_lookup(sheet)
    cell_group_list = lookup.keys()
    for cell_group in cell_group_list:
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
        sheet.unmerge_cells(str(cell_group))
        for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
            for cell in row:
                cell.value = lookup[cell_group]
    wbook.save(k.iloc[0]['Device'] + '_' + k.iloc[0]['Process'] + '.xlsx')


# # Tool change & Spanker Cleaning & Mold cleaning

# In[14]:


com = {
    'a': ['LSD001', 'LSD002', 'LSD008', 'LSD012', 'LSD022', 'LSD023', 'LSD027', 'LSD035', 'LSD039', 'LSD040', 'LSD042', 'LSD003', 'LSD004', 'LSD005', 'LSD006', 'LSD007', 'LSD009', 'LSD010', 'LSD013', 'LSD014', 'LSD015', 'LSD016', 'LSD018', 'LSD019', 'LSD024', 'LSD025', 'LSD026', 'LSD036', 'LSD037', 'LSD041'],
    'b': ['LMD002','LMD003','LMD004','LMD005','LMD006','LMD007','LMD009','LMD010','LMD011','LMD012','LMD013','LMD014']
      }


# In[15]:


for x,y in com.items():
    if x == 'a':
        filter1 = df['EQ_ID'].isin(y)
        df2 = df[filter1]
        k = pd.crosstab([df2.LEVEL6_NAME],df2.Month,values=df2.DURATION_HOUR,aggfunc=sum,rownames=['Level_5'],colnames=['date'], normalize='columns')
        display(k.loc[['CHANGE COLLET','WAITING (TOOL CHANGE OP)']])
    else:
        filter1 = df['EQ_ID'].isin(y)
        df2 = df[filter1]
        k = pd.crosstab([df2.LEVEL6_NAME],df2.Month,values=df2.DURATION_HOUR,aggfunc=sum,rownames=['Level_5'],colnames=['date'], normalize='columns')
        display(k.loc[['IRREGULAR MOLD CLEANING','REGULAR MOLD CLEANING','5S']])


# In[ ]:





# In[66]:


def create_merged_cell_lookup(sheet) -> dict:
    merged_lookup = {}
    for cell_group in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
        if min_col == max_col:
            top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
            merged_lookup[str(cell_group)] = top_left_cell_value
    return merged_lookup


# In[67]:


wbook = openpyxl.load_workbook('test.xlsx')
sheet = wbook["Sheet1"]
lookup = create_merged_cell_lookup(sheet)
cell_group_list = lookup.keys()
for cell_group in cell_group_list:
    min_col, min_row, max_col, max_row = range_boundaries(str(cell_group))
    sheet.unmerge_cells(str(cell_group))
    for row in sheet.iter_rows(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row):
        for cell in row:
            cell.value = lookup[cell_group]
wbook.save('test.xlsx')


# In[ ]:




